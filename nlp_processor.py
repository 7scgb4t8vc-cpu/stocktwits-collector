"""
NLP Sentiment Processor
========================
Reads stocktwits.csv, cleans message text, then sends ALL messages
in a single batched Claude API call to classify sentiment.

Outputs:
  - data/nlp_output.csv     : per-message label, score, clean text (read by the Flask dashboard)
  - data/sentiment_scores.csv : per-stock aggregate sentiment score (signed avg of nlp_score)
  - "NLP Sentiment" sheet in stocktwits_data.xlsx

Each message gets:
  - nlp_label  : bullish | bearish | neutral | mixed
  - nlp_score  : float 0.0–1.0 (confidence/strength)
  - clean_text : cleaned version of the original message
"""

import csv
import json
import re
import os
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

DATA_CSV       = Path("data/stocktwits.csv")
NLP_CSV        = Path("data/nlp_output.csv")
SENTIMENT_CSV  = Path("data/sentiment_scores.csv")
EXCEL_FILE     = Path("data/stocktwits_data.xlsx")
MODEL      = "claude-haiku-4-5"   # fast + cheap; ideal for bulk classification
MAX_TOKENS = 8000
BATCH_SIZE = 500                   # max messages per API call (context window safety)

NLP_FIELDS = ["timestamp", "symbol", "clean_text", "original_sentiment", "nlp_label", "nlp_score"]
SENTIMENT_FIELDS = ["symbol", "sentiment_score", "bullish_count", "bearish_count", "neutral_count", "mixed_count", "total_messages"]

# ── Text cleaning ─────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    """
    Clean a raw StockTwits message:
    - Remove URLs
    - Remove ticker cashtags (e.g. $NVDA)
    - Remove emojis and non-ASCII symbols
    - Remove excess whitespace
    - Lowercase
    """
    # Remove URLs
    text = re.sub(r"http\S+|www\.\S+", "", text)
    # Remove cashtags
    text = re.sub(r"\$[A-Z]{1,5}", "", text)
    # Remove mentions
    text = re.sub(r"@\w+", "", text)
    # Remove emojis and non-ASCII
    text = text.encode("ascii", "ignore").decode("ascii")
    # Remove special characters except basic punctuation
    text = re.sub(r"[^a-zA-Z0-9\s\.\,\!\?\'\-]", "", text)
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    return text.lower()


# ── Batch sentiment classification ───────────────────────────────────────────

def classify_batch(messages: list[dict], client: anthropic.Anthropic) -> dict:
    """
    Send a batch of messages to Claude in one API call.
    Each message dict has: {id, symbol, text}
    Returns a dict mapping id -> {label, score}
    """
    prompt = f"""You are a financial sentiment classifier for stock market social media posts.

Classify the sentiment of each message below. For each one, return:
- id: the exact id provided
- label: one of "bullish", "bearish", "neutral", or "mixed"
- score: a float from 0.0 to 1.0 representing confidence/strength of sentiment
  (1.0 = very strong clear sentiment, 0.5 = moderate, 0.1 = very weak/ambiguous)

Rules:
- "bullish" = positive outlook, expecting price to rise
- "bearish" = negative outlook, expecting price to fall
- "mixed" = contains both bullish and bearish signals
- "neutral" = no clear directional sentiment (news, questions, general commentary)
- Base your score on how strongly and clearly the sentiment is expressed
- Ignore messages that are too short or meaningless — classify as "neutral" with score 0.1

Messages to classify:
{json.dumps(messages, indent=2, ensure_ascii=False)}

Return ONLY a valid JSON array. No explanation, no markdown, no extra text.
Format: [{{"id": 1, "label": "bullish", "score": 0.85}}, ...]"""

    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        messages=[{"role": "user", "content": prompt}]
    )

    raw = response.content[0].text.strip()
    raw = raw.encode("utf-8", "ignore").decode("utf-8")

    # Strip markdown fences if present
    raw = re.sub(r"^```json\s*|^```\s*|```$", "", raw, flags=re.MULTILINE).strip()

    results = json.loads(raw)
    return {r["id"]: {"label": r["label"], "score": r["score"]} for r in results}


# ── Per-stock sentiment score ─────────────────────────────────────────────────

def compute_sentiment_scores(nlp_rows: list[dict]) -> list[dict]:
    """
    Aggregate per-message nlp_label + nlp_score into a single signed
    sentiment score per symbol, in the range -1.0 (very bearish) to
    +1.0 (very bullish).

    Signing rule per message:
      bullish ->  +score
      bearish ->  -score
      mixed   ->  +/-(score / 2)   (counted as half-weight in both directions, net ~0 unless skewed)
      neutral ->   0
    Per-stock score = average of signed values across all its messages.
    """
    by_symbol = {}
    for row in nlp_rows:
        sym = row.get("symbol", "")
        if not sym:
            continue
        by_symbol.setdefault(sym, []).append(row)

    output = []
    for sym, rows in by_symbol.items():
        signed_total = 0.0
        counts = {"bullish": 0, "bearish": 0, "neutral": 0, "mixed": 0}

        for row in rows:
            label = row.get("nlp_label", "neutral")
            score = float(row.get("nlp_score", 0.0) or 0.0)
            counts[label] = counts.get(label, 0) + 1

            if label == "bullish":
                signed_total += score
            elif label == "bearish":
                signed_total -= score
            elif label == "mixed":
                signed_total += 0.0  # net neutral contribution; strength still counted via 'mixed_count'
            # neutral contributes 0

        total = len(rows)
        avg_signed = signed_total / total if total else 0.0
        avg_signed = max(-1.0, min(1.0, avg_signed))  # clamp just in case

        output.append({
            "symbol":          sym,
            "sentiment_score": round(avg_signed, 4),
            "bullish_count":   counts["bullish"],
            "bearish_count":   counts["bearish"],
            "neutral_count":   counts["neutral"],
            "mixed_count":     counts["mixed"],
            "total_messages":  total,
        })

    return sorted(output, key=lambda r: r["sentiment_score"], reverse=True)


# ── CSV export ─────────────────────────────────────────────────────────────────

def write_nlp_csv(nlp_rows: list[dict]):
    NLP_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(NLP_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=NLP_FIELDS)
        writer.writeheader()
        writer.writerows(nlp_rows)
    print(f"  Saved {NLP_CSV} — {len(nlp_rows)} rows.")


def write_sentiment_csv(sentiment_rows: list[dict]):
    SENTIMENT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(SENTIMENT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=SENTIMENT_FIELDS)
        writer.writeheader()
        writer.writerows(sentiment_rows)
    print(f"  Saved {SENTIMENT_CSV} — {len(sentiment_rows)} symbols.")


# ── Excel export ──────────────────────────────────────────────────────────────

def write_nlp_sheet(nlp_rows: list[dict]):
    """Add or replace the 'NLP Sentiment' sheet in the existing Excel workbook."""
    if not EXCEL_FILE.exists():
        print(f"  ✗ {EXCEL_FILE} not found — run the collector first.")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Remove existing NLP sheet if present
    if "NLP Sentiment" in wb.sheetnames:
        del wb["NLP Sentiment"]

    ws = wb.create_sheet("NLP Sentiment")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    label_colors = {
        "bullish": "C6EFCE",   # green
        "bearish": "FFC7CE",   # red
        "mixed":   "FFEB9C",   # yellow
        "neutral": "FFFFFF",   # white
    }

    # Header row
    headers = ["TIMESTAMP", "SYMBOL", "CLEAN TEXT", "ORIGINAL SENTIMENT", "NLP LABEL", "NLP SCORE"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for r, row in enumerate(nlp_rows, 2):
        label = row.get("nlp_label", "neutral")
        fill  = PatternFill("solid", fgColor=label_colors.get(label, "FFFFFF"))
        for col, key in enumerate(NLP_FIELDS, 1):
            cell       = ws.cell(row=r, column=col, value=row.get(key, ""))
            cell.fill  = fill

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    wb.save(EXCEL_FILE)
    print(f"  Saved 'NLP Sentiment' sheet → {EXCEL_FILE}")


def write_sentiment_sheet(sentiment_rows: list[dict]):
    """Add or replace the 'Sentiment Scores' sheet in the existing Excel workbook."""
    if not EXCEL_FILE.exists():
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)

    if "Sentiment Scores" in wb.sheetnames:
        del wb["Sentiment Scores"]

    ws = wb.create_sheet("Sentiment Scores")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    headers = ["SYMBOL", "SENTIMENT SCORE", "BULLISH", "BEARISH", "NEUTRAL", "MIXED", "TOTAL MESSAGES"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(sentiment_rows, 2):
        score = row["sentiment_score"]
        if score > 0.15:
            fill_color = "C6EFCE"  # green
        elif score < -0.15:
            fill_color = "FFC7CE"  # red
        else:
            fill_color = "FFFFFF"  # neutral
        fill = PatternFill("solid", fgColor=fill_color)
        for col, key in enumerate(SENTIMENT_FIELDS, 1):
            cell = ws.cell(row=r, column=col, value=row.get(key, ""))
            cell.fill = fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 16

    wb.save(EXCEL_FILE)
    print(f"  Saved 'Sentiment Scores' sheet → {EXCEL_FILE}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*55}")
    print("NLP Sentiment Processor")
    print(f"{'='*55}")

    # Load API key from environment
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("✗ ANTHROPIC_API_KEY environment variable not set.")
        return

    client = anthropic.Anthropic(api_key=api_key)

    # Load messages from CSV
    if not DATA_CSV.exists():
        print(f"✗ {DATA_CSV} not found — run the collector first.")
        return

    with open(DATA_CSV, "r", encoding="utf-8") as f:
        raw_rows = list(csv.DictReader(f))

    print(f"\n  Loaded {len(raw_rows)} messages from {DATA_CSV}")

    # Clean text
    print("  Cleaning message text...")
    cleaned = []
    for i, row in enumerate(raw_rows):
        cleaned.append({
            "id":                i,
            "timestamp":         row.get("timestamp", ""),
            "symbol":            row.get("symbol", ""),
            "original_sentiment": row.get("sentiment", "None"),
            "clean_text":        clean_text(row.get("message", "")),
        })

    # Remove duplicate posts (same clean_text seen more than once)
    print("  Removing duplicate posts...")
    seen_texts = set()
    deduped    = []
    for c in cleaned:
        if c["clean_text"] not in seen_texts:
            seen_texts.add(c["clean_text"])
            deduped.append(c)
    duplicates_removed = len(cleaned) - len(deduped)
    print(f"  {duplicates_removed} duplicate(s) removed, {len(deduped)} unique messages remaining.")

    # Filter out empty messages after cleaning
    to_classify = [c for c in deduped if len(c["clean_text"]) > 5]
    skipped     = len(deduped) - len(to_classify)
    print(f"  {len(to_classify)} messages to classify, {skipped} skipped (too short after cleaning).")

    # Batch into chunks of BATCH_SIZE
    results = {}
    chunks  = [to_classify[i:i+BATCH_SIZE] for i in range(0, len(to_classify), BATCH_SIZE)]
    print(f"\n  Sending {len(chunks)} batch(es) to Claude ({MODEL})...")

    for batch_num, chunk in enumerate(chunks, 1):
        print(f"  Batch {batch_num}/{len(chunks)} — {len(chunk)} messages...", end=" ")
        batch_input = [{"id": c["id"], "symbol": c["symbol"], "text": c["clean_text"]} for c in chunk]
        try:
            batch_results = classify_batch(batch_input, client)
            results.update(batch_results)
            print("✓")
        except Exception as e:
            safe_err = str(e).encode("ascii", "backslashreplace").decode("ascii")
            print(f"✗ Error: {safe_err}")

    # Assemble final rows
    nlp_rows = []
    for c in cleaned:
        result = results.get(c["id"], {"label": "neutral", "score": 0.0})
        nlp_rows.append({
            "timestamp":          c["timestamp"],
            "symbol":             c["symbol"],
            "clean_text":         c["clean_text"],
            "original_sentiment": c["original_sentiment"],
            "nlp_label":          result["label"],
            "nlp_score":          result["score"],
        })

    # Summary stats
    label_counts = {}
    for row in nlp_rows:
        label_counts[row["nlp_label"]] = label_counts.get(row["nlp_label"], 0) + 1
    print(f"\n  Sentiment breakdown:")
    for label, count in sorted(label_counts.items()):
        print(f"    {label:10s}: {count}")

    # Per-stock sentiment score
    print("\n  Computing per-stock sentiment scores...")
    sentiment_rows = compute_sentiment_scores(nlp_rows)
    for row in sentiment_rows:
        print(f"    {row['symbol']:6s} score={row['sentiment_score']:+.3f}  "
              f"(bull={row['bullish_count']} bear={row['bearish_count']} "
              f"neu={row['neutral_count']} mix={row['mixed_count']})")

    # Write CSVs (these are what the Flask dashboard reads)
    print("\n  Writing CSV outputs...")
    write_nlp_csv(nlp_rows)
    write_sentiment_csv(sentiment_rows)

    # Write to Excel
    print("\n  Writing Excel sheets...")
    write_nlp_sheet(nlp_rows)
    write_sentiment_sheet(sentiment_rows)

    print("\n✓ NLP processing complete!")


if __name__ == "__main__":
    main()
