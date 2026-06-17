"""
StockTwits + FinViz Elite Collector
=====================================
Each run:
1. Fetches pre-filtered stocks from FinViz Elite screener export
   (filters: volume > 100k, rel volume > 1, change up — applied on FinViz side)
2. Takes the first 7 valid stocks (excludes ETFs via market cap check)
3. Collects StockTwits messages for those stocks
4. Upserts FinViz data (one row per symbol, always current)
5. Updates frequency.csv with total mention counts per symbol
6. Exports everything to a formatted Excel workbook

Runs in under 60 seconds.
"""

import json
import time
import csv
import io
import os
import re
from datetime import datetime
import pytz
from pathlib import Path

from curl_cffi import requests as curl_requests
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ────────────────────────────────────────────────────────────────────

BASE_URL      = "https://api.stocktwits.com/api/2"
IMPERSONATE   = "chrome120"
REQUEST_DELAY = 1.0
TOP_N_KEEP    = 7
MAX_NEW_MSGS  = 10
DATA_CSV      = Path("data/stocktwits.csv")
FREQ_CSV      = Path("data/frequency.csv")
FINVIZ_CSV    = Path("data/finviz.csv")
EXCEL_FILE    = Path("data/stocktwits_data.xlsx")
CURSOR_FILE   = Path("data/cursors.json")

# FinViz Elite screener export URL (filters applied on FinViz side)
FINVIZ_OVERVIEW_URL  = "https://elite.finviz.com/export?v=171&f=sh_curvol_o100,sh_relvol_o2,ta_change_u&ft=4&auth={token}"
FINVIZ_TECHNICAL_URL = "https://elite.finviz.com/export?v=171&f=sh_curvol_o100,sh_relvol_o2,ta_change_u&ft=4&auth={token}" 

ST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":         "https://stocktwits.com/",
    "Origin":          "https://stocktwits.com",
}

CSV_FIELDS    = ["timestamp", "symbol", "message", "sentiment"]
FINVIZ_FIELDS = [
    "timestamp", "symbol", "price", "change_pct", "volume", "avg_volume",
    "rel_volume", "market_cap", "rsi", "beta", "52w_high", "52w_low",
    "sector", "industry"
]

# ── Cursor tracking ───────────────────────────────────────────────────────────

def load_cursors() -> dict:
    if CURSOR_FILE.exists():
        with open(CURSOR_FILE, "r") as f:
            return json.load(f)
    return {}


def save_cursors(cursors: dict):
    CURSOR_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(CURSOR_FILE, "w") as f:
        json.dump(cursors, f, indent=2)


# ── FinViz Elite screener ─────────────────────────────────────────────────────

def fetch_finviz_screener(token: str) -> list[dict]:
    """
    Fetch pre-filtered stocks from FinViz Elite screener export.
    Returns a list of dicts with FinViz columns.
    Filters applied on FinViz side: volume > 100k, rel vol > 1, change up.
    """
    url  = FINVIZ_EXPORT_URL.format(token=token)
    print(f"  Fetching URL: {url[:80]}...")
    try:
        resp = curl_requests.get(url, impersonate=IMPERSONATE, timeout=20)
        print(f"  HTTP status: {resp.status_code}")
        resp.raise_for_status()
    except Exception as e:
        print(f"  Error fetching FinViz screener: {e}")
        print(f"  Response text: {resp.text[:200] if resp else 'no response'}")
        return []

    reader = csv.DictReader(io.StringIO(resp.text))
    rows = list(reader)
    if rows:
        print(f"  FinViz columns: {list(rows[0].keys())}")
    return rows


def parse_finviz_row(row: dict) -> dict:
    """Map FinViz Elite CSV columns to our internal field names."""
    return {
        "price":      row.get("Price",           ""),
        "change_pct": row.get("Change",          ""),
        "volume":     row.get("Volume",          ""),
        "avg_volume": row.get("Avg Volume",  ""),
        "rel_volume": row.get("Rel Volume", ""),
        "market_cap": row.get("Market Cap",      ""),
        "rsi":        row.get("RSI (14)",        ""),
        "beta":       row.get("Beta",            ""),
        "52w_high":   row.get("52W High",        ""),
        "52w_low":    row.get("52W Low",         ""),
        "sector":     row.get("Sector",          ""),
        "industry":   row.get("Industry",        ""),
    }


# ── StockTwits fetchers ───────────────────────────────────────────────────────

def fetch_trending() -> list:
    url  = f"{BASE_URL}/trending/symbols.json"
    resp = curl_requests.get(url, headers=ST_HEADERS, impersonate=IMPERSONATE, timeout=20)
    resp.raise_for_status()
    return resp.json().get("symbols", [])


def fetch_new_messages(symbol: str, since_id) -> list:
    url    = f"{BASE_URL}/streams/symbol/{symbol}.json"
    params = {"limit": MAX_NEW_MSGS}
    if since_id:
        params["since"] = since_id
    resp = curl_requests.get(url, params=params, headers=ST_HEADERS, impersonate=IMPERSONATE, timeout=20)
    resp.raise_for_status()
    return resp.json().get("messages", [])


def get_sentiment(msg: dict) -> str:
    entities = msg.get("entities", {})
    if entities.get("sentiment"):
        return entities["sentiment"].get("basic", "None") or "None"
    return "None"


# ── Output helpers ────────────────────────────────────────────────────────────

def append_to_csv(rows: list, path: Path, fields: list):
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists()
    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        if write_header:
            writer.writeheader()
        writer.writerows(rows)


def upsert_finviz_csv(new_rows: list, path: Path, fields: list):
    """Update existing rows by symbol, insert if new. One row per symbol."""
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = {}
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existing[row["symbol"]] = row
    for row in new_rows:
        existing[row["symbol"]] = row
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(existing.values())


# ── Frequency tracking ────────────────────────────────────────────────────────

def update_frequency():
    if not DATA_CSV.exists():
        return

    counts    = {}
    last_seen = {}

    with open(DATA_CSV, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            s = row["symbol"]
            counts[s]    = counts.get(s, 0) + 1
            last_seen[s] = row["timestamp"]

    FREQ_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(FREQ_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["symbol", "mention_count", "last_seen"])
        writer.writeheader()
        for s, c in sorted(counts.items(), key=lambda x: x[1], reverse=True):
            writer.writerow({"symbol": s, "mention_count": c, "last_seen": last_seen[s]})

    print(f"  Updated {FREQ_CSV} — {len(counts)} symbols tracked.")


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel():
    wb          = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    def style_header(cell):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Raw Messages ──
    ws1 = wb.active
    ws1.title = "Raw Messages"
    sentiment_colors = {"Bullish": "C6EFCE", "Bearish": "FFC7CE", "None": "FFFFFF"}

    if DATA_CSV.exists():
        with open(DATA_CSV, "r", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        for col, h in enumerate(CSV_FIELDS, 1):
            style_header(ws1.cell(row=1, column=col, value=h.upper()))
        for r, row in enumerate(rows, 2):
            fill = PatternFill("solid", fgColor=sentiment_colors.get(row.get("sentiment", "None"), "FFFFFF"))
            for col, key in enumerate(CSV_FIELDS, 1):
                ws1.cell(row=r, column=col, value=row.get(key, "")).fill = fill
        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 10
        ws1.column_dimensions["C"].width = 80
        ws1.column_dimensions["D"].width = 12

    # ── Sheet 2: Ticker Frequency ──
    ws2       = wb.create_sheet("Ticker Frequency")
    freq_cols = ["symbol", "mention_count", "last_seen"]

    if FREQ_CSV.exists():
        with open(FREQ_CSV, "r", encoding="utf-8") as f:
            freq_rows = list(csv.DictReader(f))
        for col, h in enumerate(freq_cols, 1):
            style_header(ws2.cell(row=1, column=col, value=h.upper()))
        for r, row in enumerate(freq_rows, 2):
            for col, key in enumerate(freq_cols, 1):
                ws2.cell(row=r, column=col, value=row.get(key, ""))
        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 22

    # ── Sheet 3: FinViz Data ──
    ws3 = wb.create_sheet("FinViz Data")

    if FINVIZ_CSV.exists():
        with open(FINVIZ_CSV, "r", encoding="utf-8") as f:
            fv_rows = list(csv.DictReader(f))
        for col, h in enumerate(FINVIZ_FIELDS, 1):
            style_header(ws3.cell(row=1, column=col, value=h.upper()))
        for r, row in enumerate(fv_rows, 2):
            for col, key in enumerate(FINVIZ_FIELDS, 1):
                ws3.cell(row=r, column=col, value=row.get(key, ""))
        col_widths = [22, 10, 10, 12, 15, 12, 14, 20]
        for i, w in enumerate(col_widths, 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)
    print(f"  Saved Excel workbook → {EXCEL_FILE}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    et        = pytz.timezone("America/New_York")
    timestamp = datetime.now(et).strftime("%Y-%m-%d %H:%M ET")
    print(f"\n{'='*55}")
    print(f"StockTwits + FinViz Elite Collector — {timestamp}")
    print(f"{'='*55}")

    # Load API token from environment
    finviz_token = os.environ.get("FINVIZ_API_TOKEN", "")
    if not finviz_token:
        print("✗ FINVIZ_API_TOKEN environment variable not set.")
        return

    cursors = load_cursors()

    # Step 1 — fetch pre-filtered stocks from FinViz Elite
    print("\nFetching FinViz Elite screener results...")
    fv_screener_rows = fetch_finviz_screener(finviz_token)

    if not fv_screener_rows:
        print("  No results from FinViz screener — filters may be too strict or market is closed.")
        return

    print(f"  {len(fv_screener_rows)} stocks passed FinViz filters.")

    # Step 2 — cross-reference with StockTwits trending, pick up to TOP_N_KEEP
    print("\nFetching StockTwits trending stocks...")
    try:
        trending      = fetch_trending()
        trending_syms = {s["symbol"] for s in trending}
    except Exception as e:
        print(f"  Error fetching trending: {e}")
        trending_syms = set()

    # Build FinViz lookup by symbol
    fv_lookup = {}
    for row in fv_screener_rows:
        sym = row.get("Ticker", "").strip()
        if sym:
            fv_lookup[sym] = row

    # Prefer stocks that are both in FinViz results AND trending on StockTwits
    # Fall back to any FinViz result if not enough overlap
    prioritized = [s for s in fv_lookup if s in trending_syms]
    remaining   = [s for s in fv_lookup if s not in trending_syms]
    candidates  = (prioritized + remaining)[:TOP_N_KEEP]

    print(f"  Trending overlap: {prioritized[:TOP_N_KEEP]}")
    print(f"  Selected ({len(candidates)}): {', '.join(candidates)}")

    st_rows = []
    fv_rows = []

    # Step 3 — collect StockTwits messages + save FinViz data
    print("\nCollecting StockTwits messages...")
    for symbol in candidates:
        since_id = cursors.get(symbol)
        fv_raw   = fv_lookup[symbol]
        fv_data  = parse_finviz_row(fv_raw, symbol)

        print(f"  [{symbol}] Price={fv_data['price']} Chg={fv_data['change_pct']} "
              f"Vol={fv_data['volume']} RelVol={fv_data['rel_volume']}")

        # Fetch StockTwits messages
        print(f"  [{symbol}] Fetching messages (since_id={since_id})...", end=" ")
        try:
            messages = fetch_new_messages(symbol, since_id)
        except Exception as e:
            print(f"✗ Error: {e}")
            messages = []

        if messages:
            cursors[symbol] = messages[0]["id"]
            for msg in messages:
                st_rows.append({
                    "timestamp": timestamp,
                    "symbol":    symbol,
                    "message":   msg.get("body", "").replace("\n", " ")[:280],
                    "sentiment": get_sentiment(msg),
                })
            print(f"{len(messages)} new messages.")
        else:
            print("No new messages.")

        fv_rows.append({"timestamp": timestamp, "symbol": symbol, **fv_data})
        time.sleep(REQUEST_DELAY)

    # Save all data
    if st_rows:
        append_to_csv(st_rows, DATA_CSV, CSV_FIELDS)
        save_cursors(cursors)
        print(f"\n✓ {len(st_rows)} new StockTwits messages appended.")
    else:
        print("\n✓ No new StockTwits messages this run.")

    if fv_rows:
        upsert_finviz_csv(fv_rows, FINVIZ_CSV, FINVIZ_FIELDS)
        print(f"✓ {len(fv_rows)} FinViz rows upserted.")

    print("\nUpdating ticker mention frequency...")
    update_frequency()

    print("\nExporting to Excel...")
    export_to_excel()

    print("\n✓ All done!")


if __name__ == "__main__":
    main()
